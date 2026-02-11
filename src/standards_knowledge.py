# src/standards_knowledge.py
"""
Comprehensive standards knowledge base extracted from source documents.
Provides full document content for AI model context.

Integrated documents:
- ISO 10535:2021 - Hoists for the transfer of persons
- ISO 80002-2:2017 - Medical device software validation
- ISO 11199-3:2005 - Walking aids (two-arm operated)
- FDA QSR/QMSR/ISO 13485:2016 Internal Audit Checklist
- ISO 3758:2023 - Textiles care labelling
- ISO 27427:2023 - Nebulizing systems
- ISO 17510:2015 - Sleep apnoea breathing therapy masks
"""

import os
import json


# ══════════════════════════════════════════════════════════════════
# DOCUMENT REGISTRY
# ══════════════════════════════════════════════════════════════════

DOCUMENT_REGISTRY = {
    "ISO 10535:2021": {
        "title": "Assistive Products - Hoists for the Transfer of Persons - Requirements and Test Methods",
        "edition": "Third edition, 2021-10 (Corrected 2023-08)",
        "ics": "11.180.10",
        "category": "Assistive Devices",
        "source_file": "ISO_10535_2021(en).pdf",
        "pages": 84,
        "text_extractable": True,
    },
    "ISO 80002-2:2017": {
        "title": "Medical Device Software - Part 2: Validation of Software for Medical Device Quality Systems",
        "edition": "First edition, 2017",
        "ics": "35.240.80",
        "category": "Software",
        "source_file": "ISO 80002-2-2017 (1).pdf",
        "pages": 94,
        "text_extractable": False,
    },
    "ISO 11199-3:2005": {
        "title": "Walking Aids Manipulated by Both Arms - Requirements and Test Methods - Part 3: Walking Tables",
        "edition": "First edition, 2005",
        "ics": "11.180.10",
        "category": "Assistive Devices",
        "source_file": "ISO 11199-3-2005.pdf",
        "pages": 30,
        "text_extractable": False,
    },
    "FDA Audit Checklist": {
        "title": "FDA QSR, QMSR & ISO 13485:2016 QMS Internal Audit Checklist",
        "edition": "Current",
        "ics": "N/A",
        "category": "Audit",
        "source_file": "FDA QSR, QMSR, ISO 13485_2016 QMS Internal Audit Checklist.xlsx",
        "pages": None,
        "text_extractable": True,
    },
    "ISO 3758:2023": {
        "title": "Textiles - Care Labelling Code Using Symbols",
        "edition": "Fourth edition, 2023 (EN ISO 3758:2024)",
        "ics": "01.080.20, 59.080.01",
        "category": "Labeling",
        "source_file": "EVS_EN_ISO_3758_2024_en.pdf",
        "pages": 40,
        "text_extractable": True,
    },
    "ISO 27427:2023": {
        "title": "Anaesthetic and Respiratory Equipment - Nebulizing Systems and Components",
        "edition": "Third edition, 2023 (EN ISO 27427:2023)",
        "ics": "11.040.10",
        "category": "Respiratory Equipment",
        "source_file": "EVS_EN_ISO_27427_2023_en.pdf",
        "pages": 44,
        "text_extractable": True,
    },
    "ISO 17510:2015": {
        "title": "Medical Devices - Sleep Apnoea Breathing Therapy - Masks and Application Accessories",
        "edition": "Second edition, 2015 (EN ISO 17510:2020)",
        "ics": "11.040.10",
        "category": "Respiratory Equipment",
        "source_file": "EVS_EN_ISO_17510;2020_en.pdf",
        "pages": 40,
        "text_extractable": True,
    },
}


# ══════════════════════════════════════════════════════════════════
# ISO 10535:2021 - HOISTS FOR TRANSFER OF PERSONS
# ══════════════════════════════════════════════════════════════════

ISO_10535_KNOWLEDGE = {
    "standard": "ISO 10535:2021",
    "title": "Assistive Products - Hoists for the Transfer of Persons - Requirements and Test Methods",
    "scope": """This document specifies requirements and test methods for hoists and body-support units intended for the transfer of persons with disability. It covers mobile hoists (floor-based), mobile hoists for standing transfer, stationary hoists (wall/ceiling/floor mounted), and body-support units (slings). It does not apply to hoists intended for emergency evacuation.""",
    "normative_references": [
        "ISO 2439 - Flexible cellular polymeric materials",
        "ISO 7176-22 - Wheelchairs - Set-up procedures",
        "ISO 10993 series - Biological evaluation of medical devices",
        "IEC 60601-1 - Medical electrical equipment - General safety",
        "IEC 60601-1-2 - EMC requirements",
        "IEC 61032 - Protection of persons against access to hazardous parts",
        "IEC 62366-1 - Usability engineering",
        "IEC 80601-2-60 - Particular requirements for respiratory high-flow therapy equipment",
    ],
    "key_terms": {
        "hoist": "device intended for the transfer of a person with disability by supporting and raising the person and positioning them",
        "body-support unit": "component of a hoist that provides direct support to the person being transferred (e.g. sling, seat, standing harness)",
        "mobile hoist": "hoist that can be moved while loaded or unloaded",
        "stationary hoist": "hoist that is permanently attached to a building structure (wall, ceiling, or floor)",
        "safe working load (SWL)": "maximum mass of the person the hoist is intended to carry under normal conditions of use",
        "spreader bar": "bar or frame from which the body-support unit is suspended",
    },
    "sections": {
        "4": {
            "title": "General Requirements and Test Methods",
            "subsections": {
                "4.1": {
                    "title": "General Requirements",
                    "content": """
4.1.1 Risk management: The manufacturer shall establish and maintain a risk management process per ISO 14971 throughout the product lifecycle. Risk management shall address risks associated with the intended use, foreseeable misuse, and failure modes of the hoist.

4.1.2 Ergonomic factors: The hoist shall be designed to minimize physical effort by the operator. Controls shall be positioned for easy access. The hoist shall allow smooth and controlled movement of the person.

4.1.3 Noise and vibration: Noise levels shall not cause alarm or distress to the person being transferred. Vibration shall not cause discomfort or injury.

4.1.4 Safety of moving and folding parts: Moving parts that could cause injury shall be guarded or designed to prevent trapping or crushing. Folding mechanisms shall have locking devices to prevent unintended collapse.

4.1.5 Prevention of traps for parts of the human body: The hoist shall be designed to prevent entrapment of fingers, limbs, or other body parts in any accessible opening. Openings between 4 mm and 25 mm shall be avoided.

4.1.6 V-shaped openings: V-shaped openings that could trap fingers or limbs shall be avoided or guarded.
"""
                },
                "4.2": {
                    "title": "General Test Methods",
                    "content": """
4.2.1 Test conditions: Testing shall be carried out at 23°C ± 5°C, relative humidity 45-75%.
4.2.2 Apparatus: Test loading devices per specified test mass ± 2%.
4.2.3 Permissible errors: Length ±1mm, Force ±2%, Angle ±1°, Mass ±2%.
4.2.4 Test report: Shall include identification of hoist, SWL, test methods used, results, date.
4.2.5 Safety and performance requirements: Static load tests at 1.5x SWL, dynamic tests at 1.35x SWL.
4.2.6 Test methods for general safety requirements: Visual inspection, dimensional checks, functional tests.
"""
                },
                "4.3": {"title": "Requirements for Body-Support Units", "content": "Body-support units shall safely support the person during transfer. Materials in contact with skin shall be tested per ISO 10993 for biocompatibility. Slings shall have clear load markings and size indicators."},
                "4.4": {"title": "Central Suspension Point", "content": "The central suspension point shall withstand static loads of at least 1.5x the SWL without permanent deformation."},
                "4.5": {"title": "Spreader Bar", "content": "Spreader bars shall be designed to prevent unintended detachment of body-support units. Attachment points shall withstand 1.5x SWL."},
                "4.6": {"title": "Performance", "content": "The hoist shall perform all intended functions (lift, lower, transfer) safely and smoothly. Lifting speed shall be controllable."},
                "4.7": {"title": "Rate of Movements", "content": "Maximum lifting/lowering speed: 100 mm/s (faster rates require additional risk analysis). Maximum horizontal movement speed: 1 m/s."},
                "4.8": {"title": "Operating Forces/Torques", "content": "Manual operating forces shall not exceed 200 N for movement and 50 N for controls."},
                "4.9": {"title": "Durability", "content": "Hoists shall complete 10,000 full lifting cycles at SWL without failure. Body-support units shall withstand 2,500 cycles of loading/unloading."},
                "4.10": {"title": "Hydraulic Components", "content": "Hydraulic systems shall have overload protection. Hoses and fittings shall withstand 2x maximum working pressure. Leak-free operation required."},
                "4.11": {"title": "Pneumatic Components", "content": "Pneumatic systems shall have pressure relief valves. Maximum operating pressure shall be clearly marked."},
                "4.12": {"title": "Machine Washable Hoists", "content": "If claimed machine-washable, the hoist shall withstand specified wash cycles without degradation of safety performance."},
                "4.13": {"title": "Information Supplied by Manufacturer", "content": "Instructions for use, labelling, and pre-sale information requirements. Must include SWL, intended user population, contraindications, maintenance schedule, and service life."},
            }
        },
        "5": {
            "title": "Mobile Hoists - Specific Requirements",
            "content": """
5.1 General requirements for mobile hoists.
5.2 Static strength: Mobile hoists shall withstand 1.5x SWL applied vertically at the lifting point without permanent deformation.
5.3 Static stability: The hoist shall not tip over when loaded at SWL with the load at the most unfavorable position. Stability factor ≥ 1.5.
5.4 Immobilizing device (brakes): Brakes shall prevent unintended movement. Parking brakes shall hold the hoist on a 3° incline at SWL.
5.5 Moving forces: Force to move the loaded hoist on a flat surface shall not exceed 200 N.
5.6 Manufacturer information: Pre-sale information including base dimensions, turning radius, and minimum door width.
"""
        },
        "6": {
            "title": "Mobile Hoists for Standing Position Transfer",
            "content": """
6.1 General requirements specific to standing transfer hoists.
6.2 Static strength: 1.5x SWL without permanent deformation.
6.3 Static stability: Stability against tipping in all directions.
6.4 Brakes shall hold on 3° incline.
6.5 Moving forces max 200 N.
6.6 Durability: 10,000 cycles.
6.7 Manufacturer information requirements.
"""
        },
        "7": {
            "title": "Stationary Hoists - Specific Requirements",
            "content": """
7.1 General requirements for stationary (fixed) hoists.
7.2 Specific safety requirements for wall/ceiling/floor mounted installations.
7.3 Static strength (free-standing): 1.5x SWL.
7.4 Static stability (free-standing): Stability against tipping.
7.5 Static strength for all other stationary hoists: Installation-specific load requirements.
7.6 Manufacturer information including installation requirements and structural load specifications.
"""
        },
        "8": {
            "title": "Pool Hoists - Specific Requirements",
            "content": "Requirements for hoists used in aquatic environments. Corrosion resistance requirements. Electrical safety requirements for wet environments. IP rating requirements."
        },
    },
    "test_methods_summary": {
        "static_load_test": "Apply 1.5x SWL for 1 minute, check for permanent deformation",
        "dynamic_load_test": "Apply 1.35x SWL during 10 lifting cycles, check function",
        "stability_test": "Load at SWL in most unfavorable position, verify no tipping",
        "durability_test": "10,000 full lift cycles at SWL, verify continued safe operation",
        "brake_test": "Verify parking brake holds on 3° incline at SWL",
        "biocompatibility": "Body contact materials tested per ISO 10993 series",
        "electrical_safety": "Powered hoists tested per IEC 60601-1 and IEC 60601-1-2",
    },
    "classification": "Class I medical device (most jurisdictions); Class IIa if powered with electronic controls (EU MDR)",
    "regulatory_notes": """
- FDA: 21 CFR 890.3480 - Powered patient transfer device (Class II, 510(k) required)
- EU MDR: Class I (manual) or Class IIa (powered) per Rule 1/Rule 9
- Harmonized with EU MDR Annex I essential requirements
- Risk management per ISO 14971 mandatory throughout lifecycle
- Usability engineering per IEC 62366-1 required
- If electrical: IEC 60601-1 compliance required
- Biocompatibility per ISO 10993 for body-contact materials
""",
}


# ══════════════════════════════════════════════════════════════════
# ISO 80002-2:2017 - MEDICAL DEVICE SOFTWARE VALIDATION
# ══════════════════════════════════════════════════════════════════

ISO_80002_2_KNOWLEDGE = {
    "standard": "ISO 80002-2:2017",
    "title": "Medical Device Software - Part 2: Validation of Software for Medical Device Quality Systems",
    "scope": """This document provides guidance on the validation of software used in the medical device quality management system (QMS). It applies to software used in the QMS, such as enterprise resource planning (ERP), manufacturing execution systems (MES), laboratory information management systems (LIMS), document management systems, and any computer system used for design, manufacturing, testing, or other quality system activities. It does NOT apply to software that is a medical device or part of a medical device (which is covered by IEC 62304).""",
    "normative_references": [
        "ISO 13485:2016 - Quality management systems for medical devices",
        "IEC 62304:2006+A1:2015 - Medical device software lifecycle",
        "GAMP 5 - Good Automated Manufacturing Practice",
        "FDA Guidance - General Principles of Software Validation (2002)",
        "21 CFR Part 11 - Electronic Records; Electronic Signatures",
    ],
    "key_concepts": {
        "computer_system_validation": "Establishing documented evidence that a computer system consistently performs according to predetermined specifications and quality attributes",
        "intended_use": "The specific purpose for which the software is used in the quality system, as documented by the organization",
        "risk_based_approach": "The level of validation effort should be commensurate with the risk associated with the use of the software in the quality system",
        "GAMP_categories": {
            "Category 1": "Infrastructure software (operating systems, databases) - minimal validation",
            "Category 3": "Non-configured products (COTS/off-the-shelf) - moderate validation",
            "Category 4": "Configured products - significant validation",
            "Category 5": "Custom applications - full validation",
        },
    },
    "sections": {
        "scope_and_application": """
This standard applies to ALL software used in the medical device quality system, including:
- ERP systems (SAP, Oracle, etc.)
- Document management systems (electronic QMS)
- Manufacturing execution systems (MES)
- Laboratory information management systems (LIMS)
- Statistical analysis software
- Computer-aided design (CAD) software
- Computer-aided manufacturing (CAM) software
- Electronic batch records
- Calibration management software
- Complaint handling and CAPA tracking systems
- Supplier management systems
- Training management systems
- Audit management systems
""",
        "validation_approach": """
The validation approach shall include:
1. Validation Planning: Define scope, approach, schedule, roles, responsibilities, and acceptance criteria
2. Risk Assessment: Determine criticality of the software based on:
   - Impact on product quality
   - Impact on patient safety
   - Impact on data integrity
   - Regulatory requirements
3. Requirement Specification: Document functional and non-functional requirements
4. Design Specification: Document how the system meets requirements (for custom/configured)
5. Installation Qualification (IQ): Verify correct installation and configuration
6. Operational Qualification (OQ): Verify system operates as intended under normal conditions
7. Performance Qualification (PQ): Verify system consistently performs as expected in the production environment
8. Validation Report: Document results, deviations, and conclusions
""",
        "ongoing_requirements": """
Validated systems require:
- Change control procedures for all modifications
- Periodic review of validation status
- Revalidation after significant changes
- Backup and disaster recovery procedures
- Security controls and access management
- Audit trails for regulated data
- Electronic signature compliance (21 CFR Part 11 if applicable)
- Training of users and administrators
""",
        "documentation_requirements": """
Required documentation includes:
- Validation Master Plan
- System Requirements Specification (SRS)
- Functional Requirements Specification (FRS)
- Design Specification (for custom/configured systems)
- Validation Protocol (IQ, OQ, PQ)
- Test scripts and test cases
- Validation Report with results summary
- Traceability matrix linking requirements to test cases
- Risk assessment documentation
- Change control records
- Periodic review reports
""",
        "regulatory_context": """
Key regulatory requirements:
- ISO 13485:2016 Clause 4.1.6: Validation of computer software used in the QMS
- ISO 13485:2016 Clause 7.5.6: Validation of processes for production and service provision
- ISO 13485:2016 Clause 7.6: Control of monitoring and measuring equipment (for software used in measurement)
- 21 CFR 820.70(i): Automated processes - validated per established protocols
- 21 CFR Part 11: Electronic records and electronic signatures
- EU MDR Annex I, Section 17: Electronic programmable systems
- FDA Guidance: General Principles of Software Validation (January 2002)
""",
    },
    "classification": "N/A - This standard applies to QMS tools, not medical devices themselves",
    "regulatory_notes": """
- Required by ISO 13485:2016 Clause 4.1.6 for all QMS software
- FDA expects compliance during inspections per 21 CFR 820.70(i)
- EU MDR expects validated QMS software as part of Notified Body audits
- GAMP 5 methodology widely accepted for medical device QMS software validation
- Risk-based approach: higher risk software requires more rigorous validation
- Common FDA 483 citation: failure to validate QMS software
""",
}


# ══════════════════════════════════════════════════════════════════
# ISO 11199-3:2005 - WALKING AIDS (TWO-ARM OPERATED)
# ══════════════════════════════════════════════════════════════════

ISO_11199_3_KNOWLEDGE = {
    "standard": "ISO 11199-3:2005",
    "title": "Walking Aids Manipulated by Both Arms - Requirements and Test Methods - Part 3: Walking Tables",
    "scope": """This part of ISO 11199 specifies requirements and test methods for walking tables that are used as walking aids manipulated by both arms. Walking tables are wheeled walking aids with a frame, forearm supports or handgrips, and typically 3 or 4 wheels. They are intended for persons who need support for both arms during walking. This standard does not apply to walking frames without wheels or rollators with only handgrips.""",
    "normative_references": [
        "ISO 11199-1:1999 - Walking aids manipulated by both arms - Part 1: Walking frames",
        "ISO 11199-2:2005 - Walking aids manipulated by both arms - Part 2: Rollators",
        "ISO 11334-1:2007 - Walking aids for one arm - Elbow crutches",
        "ISO 14971 - Risk management for medical devices",
        "ISO 7176 series - Wheelchairs",
    ],
    "key_terms": {
        "walking table": "Walking aid with a frame structure, forearm supports and/or handgrips, and wheels, intended to be manipulated by both arms to provide support during walking",
        "forearm support": "Component of the walking table on which the user rests their forearms during use",
        "handgrip": "Component of the walking table designed to be gripped by the user's hand",
        "safe working load": "Maximum user mass for which the walking table is designed",
    },
    "sections": {
        "requirements": """
Key requirements include:
1. Static Strength: Walking tables shall withstand 1.5x SWL applied vertically without permanent deformation
2. Stability: Walking tables shall not tip in any direction under normal use conditions. Side stability and forward/rearward stability tests required
3. Fatigue Strength: 200,000 cycles of loading at SWL without failure
4. Braking: Brakes (if fitted) shall hold the walking table stationary on a 5° incline with user mass applied
5. Rolling Resistance: Force to move walking table on flat surface shall not exceed 30 N unloaded
6. Handgrip Strength: Handgrips shall withstand 1,000 N force without failure
7. Forearm Support: Forearm supports shall withstand 500 N vertical load without permanent deformation
8. Sharp Edges: No sharp edges or protrusions accessible to the user
9. Folding Mechanism: If foldable, shall have positive locking mechanism to prevent unintended collapse
10. Wheel Requirements: Wheels shall be suitable for intended environment; casters shall swivel freely
""",
        "test_methods": """
Test methods include:
- Static load test: Apply vertical force at user support points
- Stability test: Tilt test in all directions
- Fatigue test: Cyclic loading at SWL for 200,000 cycles
- Brake test: Incline test with specified load
- Rolling resistance: Force measurement on flat, smooth surface
- Impact test: Drop test from specified height
- Durability of adjustments: Cycling of height and angle adjustments
- Wheel swivel test: Verify free rotation of casters
""",
        "labelling_requirements": """
The manufacturer shall provide:
- Maximum user mass (SWL) clearly marked on the walking table
- User instructions in the language of the intended market
- Assembly instructions (if required)
- Maintenance instructions including inspection intervals
- Warnings about hazardous conditions (stairs, slopes, wet surfaces)
- Information about height adjustment range
- Contraindications for use
""",
    },
    "classification": "Class I medical device (most jurisdictions)",
    "regulatory_notes": """
- FDA: 21 CFR 890.3420 - Walking aid (Class I, 510(k) exempt for most designs)
- EU MDR: Class I per Rule 1 (non-invasive, non-powered)
- Harmonized standard under EU MDR for walking aids
- Risk management per ISO 14971 required
- Biocompatibility per ISO 10993 for body-contact materials (forearm supports, handgrips)
- Part of ISO 11199 series covering all two-arm walking aids
""",
}


# ══════════════════════════════════════════════════════════════════
# FDA QSR/QMSR/ISO 13485 INTERNAL AUDIT CHECKLIST
# ══════════════════════════════════════════════════════════════════

FDA_AUDIT_CHECKLIST = {
    "title": "FDA QSR, QMSR & ISO 13485:2016 QMS Internal Audit Checklist",
    "description": "Comprehensive internal audit checklist mapping ISO 13485:2016, FDA 21 CFR Part 820 QSR (1996), and FDA QMSR (2026) requirements. Contains 169 audit items organized by QMS subsystem.",
    "rating_system": {
        "NC": "Non-Conformance",
        "OFI": "Opportunity for Improvement",
        "PP": "Positive Practice",
        "A": "Acceptable",
    },
    "subsystems": {
        "Management Controls": {
            "items": [
                {"item": 1, "detail": "Ensure Quality Manual defines scope of QMS, procedures (or reference to) within QMS, and description of the interaction of processes within QMS", "iso_ref": "ISO 13485:2016: 4.1, 4.2.2", "qsr_ref": "", "qmsr_ref": "ISO 13485:2016: 4.1, 4.2.2", "auditor_notes": "Review quality manual; Review QMS metrics; Review critical processes and procedures"},
                {"item": 2, "detail": "Verify criteria and methods are in place to monitor and control processes for effectiveness", "iso_ref": "ISO 13485:2016: 4.1.3(a), 4.2.1(d), 8.4", "qsr_ref": "", "qmsr_ref": "ISO 13485:2016: 4.1.3(a), 4.2.1(d), 8.4", "auditor_notes": "Review QMS metrics; review management reviews"},
                {"item": 3, "detail": "Verify firm has established and conducts Management Reviews, at least annually", "iso_ref": "ISO 13485:2016: 5.1(d), 5.6", "qsr_ref": "21 CFR 820.5, 820.20(c)", "qmsr_ref": "ISO 13485:2016: 5.1(d), 5.6", "auditor_notes": "Request procedure in advance; review management reviews"},
                {"item": 4, "detail": "Confirm management reviews examine suitability and effectiveness of quality systems, improvements needed because of customer requirements, and resource needs", "iso_ref": "ISO 13485:2016: 4.1.3(c), 5.6.1, 5.6.3, 6.1, 8.4", "qsr_ref": "21 CFR 820.20(c)", "qmsr_ref": "ISO 13485:2016: 4.1.3(c), 5.6.1, 5.6.3, 6.1, 8.4", "auditor_notes": "Review procedure"},
                {"item": 5, "detail": "Ensure management review addresses audit results, customer feedback, process performance, CAPAs, previous management reviews, changes to QMS, recommendations for improvement, and new or revised regulatory requirements", "iso_ref": "ISO 13485:2016: 5.6.2", "qsr_ref": "", "qmsr_ref": "ISO 13485:2016: 5.6.2", "auditor_notes": "Review management reviews"},
                {"item": 6, "detail": "Verify firm has established a Quality Manual and Quality System Procedures and Instructions that are appropriate", "iso_ref": "ISO 13485:2016: 4.1.2(a), 4.2.1(b), (c)", "qsr_ref": "21 CFR 820.5, 820.20(c), (d), (e), 820.22", "qmsr_ref": "ISO 13485:2016: 4.1.2(a), 4.2.1(b), (c)", "auditor_notes": "Request quality manual and procedures in advance; review documents"},
                {"item": 7, "detail": "Verify firm has established Quality Plan", "iso_ref": "ISO 13485:2016: 4.2.1(d), 5.4", "qsr_ref": "21 CFR 820.20(d)", "qmsr_ref": "ISO 13485:2016: 4.2.1(d), 5.4", "auditor_notes": "Request quality plan in advance"},
                {"item": 8, "detail": "Confirm that Quality Planning addresses QMS needs and Quality Objectives", "iso_ref": "ISO 13485:2016: 5.4.2", "qsr_ref": "21 CFR 820.20(a), (d)", "qmsr_ref": "ISO 13485:2016: 5.4.2", "auditor_notes": "Review quality plan"},
                {"item": 9, "detail": "Verify firm has implemented Quality Policy and Quality Objectives", "iso_ref": "ISO 13485:2016: 4.2.1(a), 5.1(b), (c), 5.3, 5.4.1", "qsr_ref": "21 CFR 820.20(a), (d)", "qmsr_ref": "ISO 13485:2016: 4.2.1(a), 5.1(b), (c), 5.3, 5.4.1", "auditor_notes": "Interview employees about quality policy; review training records"},
                {"item": 10, "detail": "Verify firm has established Quality Audit procedures and conducts audits", "iso_ref": "ISO 13485:2016: 4.2, 8.2.4", "qsr_ref": "21 CFR 820.20(c), 820.22", "qmsr_ref": "ISO 13485:2016: 4.2, 8.2.4", "auditor_notes": "Request procedure in advance; review audit schedule and documents; review auditor training"},
                {"item": 11, "detail": "Ensure quality audits examine compliance and effectiveness", "iso_ref": "ISO 13485:2016: 4.1.3(c), 4.2.1(d), 8.2.4", "qsr_ref": "21 CFR 820.22", "qmsr_ref": "ISO 13485:2016: 4.1.3(c), 4.2.1(d), 8.2.4", "auditor_notes": "Review procedure; review audit records"},
                {"item": 12, "detail": "Verify that auditors are trained", "iso_ref": "ISO 13485:2016: 6.2, 8.2.4", "qsr_ref": "21 CFR 820.22", "qmsr_ref": "ISO 13485:2016: 6.2, 8.2.4", "auditor_notes": "Review audit records; review training records"},
                {"item": 13, "detail": "Ensure that audits are conducted by objective parties", "iso_ref": "ISO 13485:2016: 8.2.4", "qsr_ref": "21 CFR 820.22", "qmsr_ref": "ISO 13485:2016: 8.2.4", "auditor_notes": "Review audit records; review training records"},
                {"item": 14, "detail": "Confirm quality audits are linked to CAPA", "iso_ref": "ISO 13485:2016: 8.2.4", "qsr_ref": "21 CFR 820.22, 820.100", "qmsr_ref": "ISO 13485:2016: 8.2.4", "auditor_notes": "Review procedures"},
                {"item": 15, "detail": "Review organizational structure of firm; confirm resources are available to support processes", "iso_ref": "ISO 13485:2016: 4.1.3(b), 5.1(e), 5.5.1, 5.5.2, 6.1, 6.2", "qsr_ref": "21 CFR 820.20(b), 820.25", "qmsr_ref": "ISO 13485:2016: 4.1.3(b), 5.1(e), 5.5.1, 5.5.2, 6.1, 6.2", "auditor_notes": "Request organizational chart(s) in advance"},
                {"item": 16, "detail": "Verify firm has defined a management representative with executive responsibility for implementing and reporting quality management system", "iso_ref": "ISO 13485:2016: 5.1, 5.5.1, 5.5.2, 6.1, 6.2", "qsr_ref": "21 CFR 820.20(b)(3), 820.25", "qmsr_ref": "ISO 13485:2016: 5.1, 5.5.1, 5.5.2, 6.1, 6.2", "auditor_notes": "Ask management representative to identify responsibility for changes to procedures, device designs, manufacturing processes; review of quality audit results; oversight and interaction with CAPA"},
                {"item": 17, "detail": "Verify appropriate responsibilities, authority, and resources are in place for quality system activities", "iso_ref": "ISO 13485:2016: 5.1(e), 5.5.1, 5.5.2, 6.1, 6.2", "qsr_ref": "21 CFR 820.5(b)(1)-(2), 820.20(b), 820.25", "qmsr_ref": "ISO 13485:2016: 5.1(e), 5.5.1, 5.5.2, 6.1, 6.2", "auditor_notes": "Interview management representative about resource allocation"},
                {"item": 18, "detail": "Verify firm has established procedures for identifying training needs; ensure personnel are trained to perform assigned responsibilities", "iso_ref": "ISO 13485:2016: 6.2", "qsr_ref": "21 CFR 820.25(b)", "qmsr_ref": "ISO 13485:2016: 6.2", "auditor_notes": "Review procedures; review training records"},
                {"item": 19, "detail": "AT AUDIT CONCLUSION: Determine if executive management ensures adequate and effective quality system is implemented. Ensure management is committed to and communicates importance of meeting customer and regulatory requirements", "iso_ref": "ISO 13485:2016: 5.1(a), 5.2, 5.5.3", "qsr_ref": "", "qmsr_ref": "ISO 13485:2016: 5.1(a), 5.2, 5.5.3", "auditor_notes": "Interview executive management; provide confirmation or failures of quality system"},
            ],
        },
        "Design & Development / Design Controls": {
            "items": [
                {"item": 1, "detail": "Verify products are subject to design controls", "iso_ref": "ISO 13485:2016: 7.1, 7.3", "qsr_ref": "21 CFR 820.30(a)", "qmsr_ref": "ISO 13485:2016: 7.1, 7.3", "auditor_notes": "Review procedure; review products"},
                {"item": 2, "detail": "Verify design control and risk management procedures are established and applied", "iso_ref": "ISO 13485:2016: 7.3", "qsr_ref": "21 CFR 820.30(a)-(j)", "qmsr_ref": "ISO 13485:2016: 7.3", "auditor_notes": "Ensure procedures address all design control elements"},
                {"item": 3, "detail": "Ensure design and development stages are identified; confirm review, verification, validation, and design transfer activities at each stage", "iso_ref": "ISO 13485:2016: 7.3.2", "qsr_ref": "21 CFR 820.30", "qmsr_ref": "ISO 13485:2016: 7.3.2", "auditor_notes": "Review procedures"},
                {"item": 4, "detail": "Select a design project for audit review", "iso_ref": "", "qsr_ref": "", "qmsr_ref": "", "auditor_notes": "Selection criteria: contains software, single product focus, risk based, result of complaints, most recent, cover product range, recent 510(k)/PMA/CE mark"},
                {"item": 5, "detail": "Review the project design & development plan, responsibilities, and interfaces", "iso_ref": "ISO 13485:2016: 7.3.2", "qsr_ref": "21 CFR 820.30(b)", "qmsr_ref": "ISO 13485:2016: 7.3.2", "auditor_notes": "Review plan milestones, phases, responsibilities, risk management, interfaces"},
                {"item": 6, "detail": "Verify design & development plan is updated, reviewed, and approved", "iso_ref": "ISO 13485:2016: 7.3.2", "qsr_ref": "21 CFR 820.30(b)", "qmsr_ref": "ISO 13485:2016: 7.3.2", "auditor_notes": "Review plan revisions; review and approval procedures"},
                {"item": 7, "detail": "Confirm design input requirements were established, reviewed, and approved; ensure customer requirements, functional, performance, safety, and statutory/regulatory requirements captured", "iso_ref": "ISO 13485:2016: 7.2.1, 7.3.3", "qsr_ref": "21 CFR 820.30(c)", "qmsr_ref": "ISO 13485:2016: 7.2.1, 7.3.3", "auditor_notes": "Review procedure; ensure requirements address intended use, functional, performance, safety, regulatory, user and patient needs"},
                {"item": 8, "detail": "Incomplete, ambiguous, and/or conflicting requirements were addressed", "iso_ref": "ISO 13485:2016: 7.3.3", "qsr_ref": "21 CFR 820.30(c)", "qmsr_ref": "ISO 13485:2016: 7.3.3", "auditor_notes": "Review procedure; review resolutions"},
                {"item": 9, "detail": "Confirm design & development outputs are established, verifiable, reviewed, and approved", "iso_ref": "ISO 13485:2016: 7.3.4(a), (c)", "qsr_ref": "21 CFR 820.30(d)", "qmsr_ref": "ISO 13485:2016: 7.3.4(a), (c)", "auditor_notes": "Review drawings, specifications, labeling, packaging, work instructions, IFUs"},
                {"item": 10, "detail": "Ensure design outputs are appropriate for purchasing, production, and servicing", "iso_ref": "ISO 13485:2016: 7.3.4(b)", "qsr_ref": "21 CFR 820.30(d)", "qmsr_ref": "ISO 13485:2016: 7.3.4(b)", "auditor_notes": "Review procedure"},
                {"item": 11, "detail": "Verify essential design outputs are identified", "iso_ref": "ISO 13485:2016: 7.3.4(d)", "qsr_ref": "21 CFR 820.30(d)", "qmsr_ref": "ISO 13485:2016: 7.3.4(d)", "auditor_notes": "Review drawings, specifications, labeling, packaging, work instructions, IFUs"},
                {"item": 12, "detail": "Confirm acceptance criteria is referenced by design outputs and was defined prior to verification and validation", "iso_ref": "ISO 13485:2016: 7.3.4(c), 7.3.6", "qsr_ref": "21 CFR 820.30(d) & (f)", "qmsr_ref": "ISO 13485:2016: 7.3.4(c), 7.3.6", "auditor_notes": "Review verification activities"},
                {"item": 13, "detail": "Determine if design verification confirmed design outputs met design input requirements", "iso_ref": "ISO 13485:2016: 7.3.6", "qsr_ref": "21 CFR 820.30(f)", "qmsr_ref": "ISO 13485:2016: 7.3.6", "auditor_notes": "Review verification activities"},
                {"item": 14, "detail": "Confirm design validation results prove device met predetermined user needs and intended uses", "iso_ref": "ISO 13485:2016: 7.3.7", "qsr_ref": "21 CFR 820.30(g)", "qmsr_ref": "ISO 13485:2016: 7.3.7", "auditor_notes": "Review validation activities"},
                {"item": 15, "detail": "Confirm design validation did not leave unresolved discrepancies", "iso_ref": "ISO 13485:2016: 7.3.7", "qsr_ref": "21 CFR 820.30(g)", "qmsr_ref": "ISO 13485:2016: 7.3.7", "auditor_notes": "Assess design and specification changes"},
                {"item": 16, "detail": "If required by regulations, confirm clinical evaluations/performance evaluations were performed", "iso_ref": "ISO 13485:2016: 7.3.7", "qsr_ref": "21 CFR 820.30(g)", "qmsr_ref": "ISO 13485:2016: 7.3.7", "auditor_notes": "Review evaluation data"},
                {"item": 17, "detail": "If device contains software, confirm software was validated", "iso_ref": "ISO 13485:2016: 7.3.2, 7.3.7", "qsr_ref": "21 CFR 820.30(g), 820.75", "qmsr_ref": "ISO 13485:2016: 7.3.2, 7.3.7", "auditor_notes": "Ensure software components satisfied design, validation, and change control requirements"},
                {"item": 18, "detail": "Determine if initial production units (or equivalents) were used for design validation", "iso_ref": "ISO 13485:2016: 7.3.7", "qsr_ref": "21 CFR 820.30(g)", "qmsr_ref": "ISO 13485:2016: 7.3.7", "auditor_notes": "Evaluate prototype/production records"},
                {"item": 19, "detail": "Confirm risk management activities were performed", "iso_ref": "ISO 13485:2016: 7.1; ISO 14971:2019", "qsr_ref": "21 CFR 820.30(g)", "qmsr_ref": "ISO 13485:2016: 7.1; ISO 14971:2019", "auditor_notes": "Review risk management file; ensure risk analysis, evaluation, and control steps are addressed"},
                {"item": 20, "detail": "Confirm design changes were controlled and validated (or verified)", "iso_ref": "ISO 13485:2016: 7.3.2, 7.3.6, 7.3.9", "qsr_ref": "21 CFR 820.30(i), 820.70(b), 820.75(c)", "qmsr_ref": "ISO 13485:2016: 7.3.2, 7.3.6, 7.3.9", "auditor_notes": "Review design changes and documentation decisions"},
                {"item": 21, "detail": "Confirm design changes have been reviewed for effect on components and product previously made", "iso_ref": "ISO 13485:2016: 7.3.2, 7.3.6, 7.3.9", "qsr_ref": "21 CFR 820.30(i), 820.70(b)", "qmsr_ref": "ISO 13485:2016: 7.3.2, 7.3.6, 7.3.9", "auditor_notes": "Review design changes and documentation decisions"},
                {"item": 22, "detail": "Determine if design reviews were conducted at appropriate stages of design & development", "iso_ref": "ISO 13485:2016: 7.2.2, 7.3.2, 7.3.5", "qsr_ref": "21 CFR 820.30(e)", "qmsr_ref": "ISO 13485:2016: 7.2.2, 7.3.2, 7.3.5", "auditor_notes": "Review design review documentation"},
                {"item": 23, "detail": "Confirm design review attendees were appropriate and included independent reviewer", "iso_ref": "ISO 13485:2016: 7.3.2, 7.3.5", "qsr_ref": "21 CFR 820.30(e)", "qmsr_ref": "ISO 13485:2016: 7.3.2, 7.3.5", "auditor_notes": "Review design review documentation"},
                {"item": 24, "detail": "Determine if design was correctly transferred to production", "iso_ref": "ISO 13485:2016: 7.3.2, 7.3.8", "qsr_ref": "21 CFR 820.30(h)", "qmsr_ref": "ISO 13485:2016: 7.3.2, 7.3.8", "auditor_notes": "Review device specifications"},
                {"item": 25, "detail": "Ensure Tech File contains design control documentation", "iso_ref": "ISO 13485:2016: 7.3.10", "qsr_ref": "21 CFR 820.30(b)-(j)", "qmsr_ref": "ISO 13485:2016: 7.3.10", "auditor_notes": "Review Technical File"},
            ],
        },
        "Corrective & Preventive Actions (CAPA)": {
            "items": [
                {"item": 1, "detail": "Verify CAPA procedures comply with regulatory requirements", "iso_ref": "ISO 13485:2016: 4.1, 4.2, 8.5", "qsr_ref": "21 CFR 820.100(a)", "qmsr_ref": "ISO 13485:2016: 4.1, 4.2, 8.5", "auditor_notes": "Review procedures"},
                {"item": 2, "detail": "Verify non-conforming product and CAPA procedures determine need for investigation and notification", "iso_ref": "ISO 13485:2016: 8.3, 8.5", "qsr_ref": "21 CFR 820.90(a), 820.100(a)(2)", "qmsr_ref": "ISO 13485:2016: 8.3, 8.5", "auditor_notes": "Review procedures"},
                {"item": 3, "detail": "Verify non-conforming product and CAPA procedures define responsibilities for review and disposition", "iso_ref": "ISO 13485:2016: 8.3, 8.5", "qsr_ref": "21 CFR 820.90(b)(1)", "qmsr_ref": "ISO 13485:2016: 8.3, 8.5", "auditor_notes": "Review procedures"},
                {"item": 4, "detail": "Ensure procedures for rework, retesting, and re-evaluation of nonconforming product exist and are followed", "iso_ref": "ISO 13485:2016: 8.3, 8.5", "qsr_ref": "21 CFR 820.90(b)(2)", "qmsr_ref": "ISO 13485:2016: 8.3, 8.5", "auditor_notes": "Review records of nonconforming products"},
                {"item": 5, "detail": "Verify appropriate records of quality problems have been created and used", "iso_ref": "ISO 13485:2016: 8.3, 8.5", "qsr_ref": "21 CFR 820.100(a)(1)", "qmsr_ref": "ISO 13485:2016: 8.3, 8.5", "auditor_notes": "Review records of acceptance activities, production test failures, returned products, service records, complaints"},
                {"item": 6, "detail": "Determine if trend analysis data indicates quality problems; determine if data used for CAPA decisions", "iso_ref": "ISO 13485:2016: 8.1, 8.2.5, 8.4, 8.5", "qsr_ref": "21 CFR 820.100(a)(1), 820.250", "qmsr_ref": "ISO 13485:2016: 8.1, 8.2.5, 8.4, 8.5", "auditor_notes": "Review records of incoming products, components, testing, SPC data"},
                {"item": 7, "detail": "Verify CAPA data is complete, accurate, and timely; compare results across multiple data sources", "iso_ref": "ISO 13485:2016: 8.4, 8.5", "qsr_ref": "21 CFR 820.100(a)(1)", "qmsr_ref": "ISO 13485:2016: 8.4, 8.5", "auditor_notes": "Review data sources; use data tables; compare results"},
                {"item": 8, "detail": "Verify appropriate statistical techniques are implemented", "iso_ref": "ISO 13485:2016: 8.1, 8.2.5, 8.4", "qsr_ref": "21 CFR 820.100(a)(1), 820.250", "qmsr_ref": "ISO 13485:2016: 8.1, 8.2.5, 8.4", "auditor_notes": "Review procedures; review techniques used"},
                {"item": 9, "detail": "Verify device failure investigations determine root cause", "iso_ref": "ISO 13485:2016: 8.3, 8.5", "qsr_ref": "21 CFR 820.100(a)(2)", "qmsr_ref": "ISO 13485:2016: 8.3, 8.5", "auditor_notes": "Review investigations"},
                {"item": 10, "detail": "Verify failure investigations are commensurate with risks", "iso_ref": "ISO 13485:2016: 8.3, 8.5", "qsr_ref": "21 CFR 820.100(a)(2), 820.90(b)", "qmsr_ref": "ISO 13485:2016: 8.3, 8.5", "auditor_notes": "Review investigations"},
                {"item": 11, "detail": "Verify controls exist to prevent non-conforming product from being released", "iso_ref": "ISO 13485:2016: 8.3", "qsr_ref": "21 CFR 820.90(b)", "qmsr_ref": "ISO 13485:2016: 8.3", "auditor_notes": "Review non-conformance records"},
                {"item": 12, "detail": "Verify appropriate actions were taken for quality problems", "iso_ref": "ISO 13485:2016: 8.2.5, 8.5.2, 8.5.3", "qsr_ref": "21 CFR 820.100(a)(3), (a)(5), (a)(4), (b)", "qmsr_ref": "ISO 13485:2016: 8.2.5, 8.5.2, 8.5.3", "auditor_notes": "Review CAPA records"},
                {"item": 13, "detail": "Determine CAPA actions were effective, verified, validated, documented, and implemented", "iso_ref": "ISO 13485:2016: 8.5", "qsr_ref": "21 CFR 820.100(a)(4), (a)(5), (b)", "qmsr_ref": "ISO 13485:2016: 8.5", "auditor_notes": "Review CAPA records"},
                {"item": 14, "detail": "Verify CAPAs and nonconformities were disseminated to responsible personnel", "iso_ref": "ISO 13485:2016: 8.3, 8.5", "qsr_ref": "21 CFR 820.100(a)(6)", "qmsr_ref": "ISO 13485:2016: 8.3, 8.5", "auditor_notes": "Review CAPA and non-conformance records"},
                {"item": 15, "detail": "Verify quality issues and CAPAs were disseminated for Management Review", "iso_ref": "ISO 13485:2016: 5.6.3, 8.3, 8.5", "qsr_ref": "21 CFR 820.100(a)(6), (a)(7)", "qmsr_ref": "ISO 13485:2016: 5.6.3, 8.3, 8.5", "auditor_notes": "Review CAPA records"},
                {"item": 16, "detail": "Verify procedures for handling complaints, investigation of advisory notices/recalls with CAPA feed-in", "iso_ref": "ISO 13485:2016: 7.2.3, 8.2.1, 8.2.2, 8.2.3", "qsr_ref": "21 CFR 820.100, 820.198", "qmsr_ref": "ISO 13485:2016: 7.2.3, 8.2.1, 8.2.2, 8.2.3", "auditor_notes": "Review procedures"},
            ],
        },
        "Production & Process Controls": {
            "description": "Covers manufacturing process controls, equipment maintenance, environmental controls, and production documentation per ISO 13485 Section 7.5 and 21 CFR 820.70-820.75.",
            "key_areas": [
                "Process validation for special processes",
                "Equipment maintenance and calibration",
                "Environmental monitoring and controls",
                "Work instructions and production records",
                "Labeling and packaging controls",
                "Process changes and revalidation",
                "Statistical techniques for process monitoring",
                "Cleanroom and contamination controls",
            ],
        },
        "Material Controls / Purchasing": {
            "description": "Covers supplier management, incoming inspection, material traceability, and purchasing controls per ISO 13485 Section 7.4 and 21 CFR 820.50.",
            "key_areas": [
                "Supplier evaluation and qualification",
                "Approved supplier list maintenance",
                "Incoming inspection procedures",
                "Material identification and traceability",
                "Supplier monitoring and re-evaluation",
                "Purchasing data and specifications",
                "Acceptance activities and records",
            ],
        },
        "Records / Document Controls": {
            "description": "Covers document control system, record management, electronic records, and retention per ISO 13485 Section 4.2 and 21 CFR 820.40/820.180/820.184.",
            "key_areas": [
                "Document approval and distribution",
                "Change control procedures",
                "Record retention and archival",
                "Electronic records and signatures (21 CFR Part 11)",
                "Device History Record (DHR)",
                "Device Master Record (DMR)",
                "Quality Management System Record (QMSR)",
            ],
        },
        "Facility & Equipment Controls": {
            "description": "Covers building/facility requirements, equipment qualification, calibration, and environmental controls per ISO 13485 Section 6.3/6.4 and 21 CFR 820.70.",
            "key_areas": [
                "Building and facility suitability",
                "Equipment qualification (IQ/OQ/PQ)",
                "Calibration program",
                "Preventive maintenance programs",
                "Environmental monitoring",
                "Utility qualification",
                "Contamination control",
            ],
        },
    },
}


# ══════════════════════════════════════════════════════════════════
# ISO 3758:2023 - TEXTILES CARE LABELLING
# ══════════════════════════════════════════════════════════════════

ISO_3758_KNOWLEDGE = {
    "standard": "ISO 3758:2023",
    "title": "Textiles - Care Labelling Code Using Symbols",
    "scope": """This document establishes a system of graphic symbols, intended for use in the marking of textile articles, and for providing information on the most severe treatment that does not cause irreversible damage to the article during the textile care process. This document applies to all textile articles in the form in which they are supplied to the end user. It is applicable to medical textiles, textile-based medical devices, and reusable textile components of medical devices (e.g., slings, straps, covers, garments).""",
    "symbol_categories": {
        "washing": {
            "description": "Washtub symbol indicating domestic washing process",
            "symbols": {
                "normal_wash_95": "Maximum wash temperature 95°C, normal mechanical action",
                "normal_wash_70": "Maximum wash temperature 70°C, normal mechanical action",
                "normal_wash_60": "Maximum wash temperature 60°C, normal mechanical action",
                "normal_wash_50": "Maximum wash temperature 50°C, normal mechanical action",
                "normal_wash_40": "Maximum wash temperature 40°C, normal mechanical action",
                "normal_wash_30": "Maximum wash temperature 30°C, normal mechanical action",
                "mild_wash": "Mild mechanical action (single underbar)",
                "very_mild_wash": "Very mild mechanical action (double underbar)",
                "hand_wash": "Hand wash only (maximum 40°C)",
                "do_not_wash": "Do not wash (crossed-out washtub)",
            },
        },
        "bleaching": {
            "description": "Triangle symbol indicating bleaching treatment",
            "symbols": {
                "any_bleach": "Any bleaching agent allowed (empty triangle)",
                "oxygen_bleach": "Only non-chlorine/oxygen bleach (triangle with two diagonal lines)",
                "do_not_bleach": "Do not bleach (crossed-out triangle)",
            },
        },
        "drying": {
            "description": "Square symbol indicating drying process",
            "symbols": {
                "tumble_dry_normal_high": "Tumble dry normal, high heat",
                "tumble_dry_normal_low": "Tumble dry normal, low heat",
                "do_not_tumble_dry": "Do not tumble dry (crossed-out circle in square)",
                "line_dry": "Line dry (vertical line in square)",
                "drip_dry": "Drip dry (three vertical lines in square)",
                "flat_dry": "Flat dry (horizontal line in square)",
                "dry_in_shade": "Dry in shade (two diagonal lines in upper left corner)",
            },
        },
        "ironing": {
            "description": "Iron symbol indicating ironing treatment",
            "symbols": {
                "iron_high": "Iron at maximum 200°C (three dots)",
                "iron_medium": "Iron at maximum 150°C (two dots)",
                "iron_low": "Iron at maximum 110°C without steam (one dot)",
                "do_not_iron": "Do not iron (crossed-out iron)",
            },
        },
        "professional_care": {
            "description": "Circle symbol indicating professional textile care",
            "symbols": {
                "dry_clean_any": "Professional dry clean, any solvent (letter A in circle)",
                "dry_clean_perc": "Professional dry clean, tetrachloroethylene and solvents listed for F (letter P in circle)",
                "dry_clean_hydrocarbon": "Professional dry clean, hydrocarbon solvents only (letter F in circle)",
                "wet_clean": "Professional wet cleaning (letter W in circle)",
                "do_not_dry_clean": "Do not dry clean (crossed-out circle)",
            },
        },
    },
    "medical_device_relevance": """
ISO 3758 is relevant to medical devices in the following contexts:
1. Reusable textile medical devices (surgical drapes, gowns, patient transfer slings, support garments)
2. Textile components of medical devices (straps, covers, padding, harnesses)
3. Labelling requirements under EU MDR Annex I Chapter III (information supplied by manufacturer)
4. ISO 15223-1 compatibility for medical device labelling
5. Reprocessing instructions for reusable medical devices per ISO 17664
6. FDA labelling requirements (21 CFR 801) for devices with textile components
7. Care instructions must be validated to ensure device performance is maintained after care treatment
""",
    "classification": "N/A - labelling standard, not a device standard",
    "regulatory_notes": """
- Referenced by medical device standards for textile components
- Essential for reusable medical device labelling compliance
- Must align with ISO 17664 (reprocessing of reusable medical devices)
- EU MDR requires care/maintenance instructions in IFU
- Relevant for slings (ISO 10535), garments, support devices with textile components
""",
}


# ══════════════════════════════════════════════════════════════════
# ISO 27427:2023 - NEBULIZING SYSTEMS
# ══════════════════════════════════════════════════════════════════

ISO_27427_KNOWLEDGE = {
    "standard": "ISO 27427:2023",
    "title": "Anaesthetic and Respiratory Equipment - Nebulizing Systems and Components",
    "scope": """This document specifies requirements for nebulizing systems and their components intended for use with humans for therapeutic or diagnostic purposes. It covers jet nebulizers, ultrasonic nebulizers, mesh/membrane nebulizers, and nebulizer systems with integrated compressors. It does not apply to drug-specific nebulizer systems that are regulated as drug-device combination products, devices for humidification only, or steam inhalers.""",
    "normative_references": [
        "ISO 5356-1 - Anaesthetic/respiratory equipment - Conical connectors",
        "ISO 7000 - Graphical symbols",
        "ISO 9360-1 - Anaesthetic/respiratory equipment - Heat and moisture exchangers",
        "ISO 10993 series - Biological evaluation of medical devices",
        "ISO 15223-1 - Symbols for medical device labelling",
        "ISO 17664 - Processing of health care products",
        "ISO 80601-2-12 - Critical care ventilators",
        "IEC 60601-1 - Medical electrical equipment - General requirements",
        "IEC 60601-1-2 - EMC requirements",
        "IEC 62304 - Medical device software lifecycle",
        "IEC 62366-1 - Usability engineering",
    ],
    "key_terms": {
        "nebulizer": "Device that generates an aerosol from a liquid for inhalation",
        "jet_nebulizer": "Nebulizer that uses compressed gas flow to generate aerosol by Bernoulli effect",
        "ultrasonic_nebulizer": "Nebulizer that uses high-frequency vibrations to generate aerosol",
        "mesh_nebulizer": "Nebulizer that forces liquid through a mesh/membrane with micropores to generate aerosol",
        "mass_median_aerodynamic_diameter": "MMAD - diameter such that 50% of the aerosol mass is contained in particles of smaller diameter",
        "fine_particle_fraction": "FPF - fraction of aerosol mass contained in particles with aerodynamic diameter < 5 micrometers",
        "nebulization_rate": "Mass of drug solution nebulized per unit time",
        "residual_volume": "Volume of drug solution remaining in the nebulizer after nebulization is complete",
    },
    "sections": {
        "general_requirements": """
- Nebulizer systems shall comply with ISO 14971 for risk management
- Materials in contact with drug solution shall be compatible and shall not adsorb drug
- Materials in contact with patient airway shall be biocompatible per ISO 10993
- Electrical safety per IEC 60601-1
- EMC per IEC 60601-1-2
- Usability per IEC 62366-1
- If containing software: IEC 62304 compliance required
""",
        "performance_requirements": """
Key performance requirements:
1. Aerosol output rate: Shall be declared by manufacturer and verified by test
2. Particle size distribution: MMAD and GSD (geometric standard deviation) shall be declared
3. Fine particle fraction (FPF): Percentage of particles < 5 μm shall be declared
4. Total drug delivered: Amount of drug available for inhalation shall be declared
5. Residual volume: Maximum residual volume shall be declared
6. Nebulization time: Time to nebulize a specified fill volume shall be declared
7. Flow resistance: Maximum inspiratory and expiratory resistance shall be specified
8. Noise level: Shall be declared; shall not exceed 70 dB(A) at 1 m
""",
        "test_methods": """
Test methods include:
- Aerosol particle size analysis by laser diffraction or cascade impaction
- Gravimetric measurement of aerosol output
- Residual volume measurement
- Flow resistance measurement per ISO 9360-1
- Noise measurement per IEC 60601-1
- Drug delivery efficiency testing with specified test solutions (NaF 0.1%)
- Durability testing (number of nebulization cycles)
- Cleaning/disinfection verification
""",
        "labelling_requirements": """
Labelling shall include:
- Intended patient population (adult, pediatric, neonatal)
- Compatible drug formulations (if applicable)
- Fill volume range (minimum and maximum)
- Recommended operating pressure/flow (for jet nebulizers)
- Cleaning and disinfection instructions
- Maximum number of reuse cycles (if reusable)
- Replacement intervals for consumable components
- Connection type and compatibility information
- Storage conditions
""",
        "reprocessing": """
For reusable nebulizers:
- Cleaning and disinfection instructions per ISO 17664
- Validated reprocessing methods
- Maximum number of reprocessing cycles
- Inspection criteria for continued use
- End-of-life indicators
""",
    },
    "classification": """
- FDA: 21 CFR 868.5630 - Nebulizer (Class II, 510(k) required, product code BYF/CBK)
- EU MDR: Class IIa per Rule 2 (non-invasive channeling/storing for administration to body)
- If drug-device combination: additional requirements per EU MDR Article 1(8) / FDA 21 CFR Part 4
""",
    "regulatory_notes": """
- Class II medical device in most jurisdictions
- FDA 510(k) pathway - predicate device comparison required
- EU MDR requires Notified Body certification
- ISO 27427:2023 is harmonized with EU MDR
- Biocompatibility per ISO 10993 required for airway-contact materials
- Electrical safety per IEC 60601-1 required for powered nebulizers
- Drug delivery performance claims require clinical evidence
- Combination product considerations if marketed with specific drug
""",
}


# ══════════════════════════════════════════════════════════════════
# ISO 17510:2015 - SLEEP APNOEA MASKS
# ══════════════════════════════════════════════════════════════════

ISO_17510_KNOWLEDGE = {
    "standard": "ISO 17510:2015",
    "title": "Medical Devices - Sleep Apnoea Breathing Therapy - Masks and Application Accessories",
    "scope": """This document specifies requirements for masks and application accessories (including headgear, chin straps, and connecting elements) intended for use in sleep apnoea breathing therapy with positive airway pressure (PAP) equipment. It applies to nasal masks, oronasal (full face) masks, nasal pillows/prongs, oral masks, and total face masks used with CPAP, APAP, and bilevel PAP devices. It does not apply to masks for ventilators, anaesthesia, or oxygen therapy (covered by other standards).""",
    "normative_references": [
        "ISO 5356-1 - Conical connectors",
        "ISO 10993 series - Biological evaluation of medical devices",
        "ISO 14971 - Risk management for medical devices",
        "ISO 17510-2 - Sleep apnoea breathing therapy equipment",
        "ISO 18562 series - Biocompatibility evaluation of breathing gas pathways",
        "ISO 80601-2-70 - PAP equipment for sleep-disordered breathing",
        "IEC 60601-1 - Medical electrical equipment - General safety",
        "IEC 62366-1 - Usability engineering",
    ],
    "key_terms": {
        "mask": "Patient interface that forms a seal on the face to deliver pressurized air from PAP equipment",
        "nasal_mask": "Mask that covers the nose only for delivery of breathing therapy",
        "oronasal_mask": "Mask that covers both the nose and mouth (full face mask)",
        "nasal_pillows": "Patient interface using small cushions that seal against the nostrils",
        "headgear": "Component that secures the mask to the patient's head",
        "anti_asphyxia_valve": "Safety valve that opens to allow ambient air breathing if the PAP device fails or is disconnected",
        "intentional_leak": "Designed opening in the mask or circuit for CO2 washout in single-limb circuits",
        "dead_space_volume": "Volume of the mask cavity between the patient's face and the exhaust/valve",
    },
    "sections": {
        "general_requirements": """
- Risk management per ISO 14971 throughout product lifecycle
- Biocompatibility per ISO 10993 for all materials contacting patient (face, airway)
- Biocompatibility of breathing gas pathway per ISO 18562
- Usability engineering per IEC 62366-1
- Materials shall not degrade with cleaning agents recommended by manufacturer
- Mask shall not restrict patient breathing if PAP device fails (anti-asphyxia requirement)
""",
        "safety_requirements": """
Critical safety requirements:
1. Anti-asphyxia protection: If the mask includes an anti-asphyxia valve, it SHALL open to ambient air if pressure drops below 2 hPa (cmH2O). CO2 rebreathing SHALL not exceed 3% at the valve opening pressure.
2. Quick-release mechanism: Mask SHALL be able to be removed from the patient in less than 5 seconds.
3. CO2 washout: For masks used with single-limb circuits, intentional leak ports SHALL provide adequate CO2 washout. CO2 concentration in inspired air SHALL not exceed 3% under worst-case conditions.
4. Dead space: Dead space volume SHALL be minimized and declared by manufacturer. For nasal masks, dead space SHALL not exceed 200 mL.
5. Resistance to breathing: Inspiratory and expiratory resistance SHALL be declared. At 50 L/min flow, resistance SHALL not exceed 2.5 hPa.
6. Noise: Noise from intentional leak SHALL be declared. At 10 hPa therapy pressure, noise SHALL not exceed 30 dB(A) at 1 m.
""",
        "performance_requirements": """
Performance requirements:
1. Seal quality: Mask SHALL maintain adequate seal at therapy pressures from 4 to 20 hPa
2. Intentional leak flow: Shall be declared at specified pressures (4, 10, 20 hPa)
3. Headgear retention: Headgear SHALL maintain mask position during therapy with the patient in all sleep positions
4. Material durability: Mask cushion SHALL maintain performance after specified number of cleaning cycles
5. Connector compatibility: Shall be compatible with ISO 5356-1 connectors (22 mm)
6. Pressure drop: Through mask system at specified flow rates
""",
        "test_methods": """
Test methods include:
- Anti-asphyxia valve function: Test at specified pressure differentials
- CO2 rebreathing: Measure CO2 concentration using breathing simulation
- Dead space volume: Water displacement method
- Resistance measurement: At specified flow rates (30, 50, 85 L/min)
- Intentional leak measurement: At specified therapy pressures
- Quick-release time: Timed removal test
- Noise measurement: Per IEC 60601-1 at specified therapy pressures
- Biocompatibility: Per ISO 10993 series for skin contact and mucosal membrane contact
- Breathing gas pathway biocompatibility: Per ISO 18562
- Durability: Repeated fitting/removal cycles; cleaning cycle testing
""",
        "labelling_requirements": """
Labelling SHALL include:
- Intended patient population (adult, pediatric)
- Mask type (nasal, oronasal, nasal pillows, total face)
- Size designation
- Anti-asphyxia valve location and function description
- Intentional leak port location
- Dead space volume
- Connection type
- Cleaning instructions with validated methods
- Replacement intervals
- Headgear sizing and adjustment instructions
- Warnings about use without PAP device
- Warnings about use with supplemental oxygen
- Compatible PAP equipment
""",
    },
    "classification": """
- FDA: 21 CFR 868.5905 - Nasal CPAP mask (Class II, 510(k) required, product code QBJ)
- EU MDR: Class IIa per Rule 2/Rule 9 (non-invasive channeling device for breathing support)
- UK MDR: Class IIa (mirrors EU classification)
- TGA (Australia): Class IIa
""",
    "regulatory_notes": """
- Class IIa medical device in most jurisdictions
- FDA 510(k) required with predicate device comparison
- EU MDR requires Notified Body certification
- ISO 17510:2015 is harmonized with EU MDR
- Biocompatibility per ISO 10993 required (skin contact - prolonged use)
- Biocompatibility of breathing gas pathway per ISO 18562
- Usability engineering critical for patient compliance (IEC 62366-1)
- Risk management per ISO 14971 mandatory
- Related standard: ISO 80601-2-70 for PAP equipment
- Post-market surveillance critical due to patient use in unsupervised home setting
""",
}


# ══════════════════════════════════════════════════════════════════
# RUNTIME PDF TEXT EXTRACTION
# ══════════════════════════════════════════════════════════════════

_pdf_cache = {}

def _get_src_dir():
    """Get the src directory path."""
    return os.path.join(os.path.dirname(os.path.abspath(__file__)))


def load_pdf_text(standard_key):
    """Load full text from a PDF file at runtime.

    Args:
        standard_key: Key from DOCUMENT_REGISTRY (e.g., "ISO 10535:2021")

    Returns:
        Full text content of the PDF, or None if not available.
    """
    if standard_key in _pdf_cache:
        return _pdf_cache[standard_key]

    doc = DOCUMENT_REGISTRY.get(standard_key)
    if not doc or not doc["text_extractable"]:
        return None

    src_dir = _get_src_dir()
    filepath = os.path.join(src_dir, doc["source_file"])

    if not os.path.exists(filepath):
        return None

    try:
        import pdfplumber
        pages = []
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text and len(text.strip()) > 20:
                    pages.append(text)
        result = "\n\n".join(pages)
        _pdf_cache[standard_key] = result
        return result
    except ImportError:
        return None
    except Exception:
        return None


def load_audit_checklist():
    """Load the FDA audit checklist from Excel at runtime.

    Returns:
        List of dicts with checklist items, or None if not available.
    """
    src_dir = _get_src_dir()
    filepath = os.path.join(src_dir, "FDA QSR, QMSR, ISO 13485_2016 QMS Internal Audit Checklist.xlsx")

    if not os.path.exists(filepath):
        return None

    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        ws = wb[wb.sheetnames[0]]
        rows = []
        for row in ws.iter_rows(min_row=7, max_row=ws.max_row, values_only=True):
            item_num = row[0]
            detail = row[1]
            iso_ref = row[2]
            qsr_ref = row[3]
            qmsr_ref = row[4]
            notes = row[5]
            if detail and str(detail).strip():
                rows.append({
                    "item": str(item_num) if item_num else "",
                    "detail": str(detail).strip(),
                    "iso_ref": str(iso_ref).strip() if iso_ref else "",
                    "qsr_ref": str(qsr_ref).strip() if qsr_ref else "",
                    "qmsr_ref": str(qmsr_ref).strip() if qmsr_ref else "",
                    "auditor_notes": str(notes).strip() if notes else "",
                })
        return rows
    except ImportError:
        return None
    except Exception:
        return None


# ══════════════════════════════════════════════════════════════════
# UNIFIED KNOWLEDGE ACCESS
# ══════════════════════════════════════════════════════════════════

ALL_STANDARDS_KNOWLEDGE = {
    "ISO 10535:2021": ISO_10535_KNOWLEDGE,
    "ISO 80002-2:2017": ISO_80002_2_KNOWLEDGE,
    "ISO 11199-3:2005": ISO_11199_3_KNOWLEDGE,
    "ISO 3758:2023": ISO_3758_KNOWLEDGE,
    "ISO 27427:2023": ISO_27427_KNOWLEDGE,
    "ISO 17510:2015": ISO_17510_KNOWLEDGE,
}


def get_standard_knowledge(standard_key):
    """Get the knowledge dict for a specific standard.

    Args:
        standard_key: Standard identifier (e.g., "ISO 10535:2021")

    Returns:
        Knowledge dict or None.
    """
    return ALL_STANDARDS_KNOWLEDGE.get(standard_key)


def search_standards_knowledge(query):
    """Search across all standards knowledge for relevant content.

    Args:
        query: Search term or phrase.

    Returns:
        List of (standard_key, section, content) tuples matching the query.
    """
    query_lower = query.lower()
    results = []

    for std_key, knowledge in ALL_STANDARDS_KNOWLEDGE.items():
        # Search in title
        if query_lower in knowledge.get("title", "").lower():
            results.append((std_key, "title", knowledge["title"]))

        # Search in scope
        if query_lower in knowledge.get("scope", "").lower():
            results.append((std_key, "scope", knowledge["scope"]))

        # Search in sections
        sections = knowledge.get("sections", {})
        for sec_key, sec_val in sections.items():
            if isinstance(sec_val, str):
                if query_lower in sec_val.lower():
                    results.append((std_key, sec_key, sec_val))
            elif isinstance(sec_val, dict):
                content = sec_val.get("content", "")
                if query_lower in content.lower():
                    results.append((std_key, f"{sec_key} - {sec_val.get('title', '')}", content))

        # Search in regulatory notes
        if query_lower in knowledge.get("regulatory_notes", "").lower():
            results.append((std_key, "regulatory_notes", knowledge["regulatory_notes"]))

    # Search audit checklist
    for subsystem, data in FDA_AUDIT_CHECKLIST.get("subsystems", {}).items():
        if isinstance(data, dict) and "items" in data:
            for item in data["items"]:
                if query_lower in item.get("detail", "").lower():
                    results.append(("FDA Audit Checklist", f"{subsystem} #{item['item']}", item["detail"]))

    return results


def get_audit_checklist_for_clause(clause):
    """Get audit checklist items relevant to a specific ISO 13485 clause.

    Args:
        clause: ISO 13485 clause number (e.g., "7.3", "8.5")

    Returns:
        List of matching audit items.
    """
    matches = []
    for subsystem, data in FDA_AUDIT_CHECKLIST.get("subsystems", {}).items():
        if isinstance(data, dict) and "items" in data:
            for item in data["items"]:
                if clause in item.get("iso_ref", "") or clause in item.get("qmsr_ref", ""):
                    matches.append({
                        "subsystem": subsystem,
                        "item": item["item"],
                        "detail": item["detail"],
                        "iso_ref": item["iso_ref"],
                        "qsr_ref": item["qsr_ref"],
                        "auditor_notes": item["auditor_notes"],
                    })
    return matches


def build_knowledge_context(query=None, standards=None):
    """Build a context string from the knowledge base for AI prompts.

    Args:
        query: Optional search query to filter relevant content.
        standards: Optional list of standard keys to include.

    Returns:
        Formatted string with relevant knowledge for AI context.
    """
    parts = []

    target_standards = standards or list(ALL_STANDARDS_KNOWLEDGE.keys())

    for std_key in target_standards:
        knowledge = ALL_STANDARDS_KNOWLEDGE.get(std_key)
        if not knowledge:
            continue

        parts.append(f"\n=== {knowledge['standard']} ===")
        parts.append(f"Title: {knowledge['title']}")
        parts.append(f"Scope: {knowledge['scope']}")

        if knowledge.get("classification"):
            parts.append(f"Classification: {knowledge['classification']}")

        if knowledge.get("regulatory_notes"):
            parts.append(f"Regulatory Notes: {knowledge['regulatory_notes']}")

        # Include key sections
        sections = knowledge.get("sections", {})
        for sec_key, sec_val in sections.items():
            if isinstance(sec_val, str):
                parts.append(f"\n{sec_key}: {sec_val}")
            elif isinstance(sec_val, dict):
                title = sec_val.get("title", sec_key)
                content = sec_val.get("content", "")
                if content:
                    parts.append(f"\n{title}: {content}")

    # Include audit checklist summary
    parts.append("\n=== FDA QSR/QMSR/ISO 13485 Internal Audit Checklist ===")
    parts.append(f"Description: {FDA_AUDIT_CHECKLIST['description']}")
    parts.append("Subsystems covered: " + ", ".join(FDA_AUDIT_CHECKLIST["subsystems"].keys()))

    return "\n".join(parts)
